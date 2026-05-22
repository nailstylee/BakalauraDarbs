from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import sys
import threading
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs

from dotenv import load_dotenv
from google import genai
from google.genai import types
from PIL import Image
from playwright.sync_api import Page, sync_playwright

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    raise RuntimeError(
        "This agent now writes an .xlsx results file and requires openpyxl. "
        "Install it with: python -m pip install openpyxl"
    ) from exc


ROOT = Path(__file__).resolve().parent
RUNS_DIR = ROOT / "runs"
IGNORED_HTML_DIRS = {".venv", "runs", "__pycache__", "node_modules", ".git"}

BASE_TASK_PAGE_BY_INDEX = {
    1: "puzzle_text.html",
    2: "puzzle_text.html",
    3: "puzzle_text.html",
    4: "puzzle_text.html",
    5: "puzzle_slide.html",
    6: "puzzle_slide.html",
    7: "puzzle_slide.html",
    8: "puzzle_slide.html",
    9: "puzzle_choose.html",
    10: "puzzle_choose.html",
    11: "puzzle_choose.html",
    12: "puzzle_choose.html",
    13: "puzzle_cover.html",
    14: "puzzle_cover.html",
    15: "puzzle_cover.html",
    16: "puzzle_cover.html",
}

PROTOTYPE_ARGUMENTS = {
    "prototype",
    "prototips",
    "shadow",
    "shadow-prototype",
    "main-prototype",
}

RESULTS_HEADERS = [
    "run_id",
    "completed_at",
    "model",
    "launch_target",
    "task_group",
    "task_type",
    "task_index",
    "variant",
    "html_file",
    "success",
    "outcome",
    "duration_seconds",
    "decision_rounds",
    "model_requests",
    "check_presses",
    "input_tokens",
    "cached_input_tokens",
    "output_tokens",
    "reasoning_tokens",
    "total_tokens",
    "estimated_cost_usd",
    "manual_error_type",
    "manual_notes",
    "run_folder",
    "result_json",
]

MANUAL_ERROR_TYPES = [
    "Instrukcija nav saprasta",
    "Attēls vai objekts interpretēts nepareizi",
    "Telpiskā spriešana vai mērķa stāvoklis nepareizs",
    "Precīza mijiedarbība neizdevās",
    "Iestrēga vai atkārtoja neefektīvu pieeju",
    "Cits",
]

TASK_TYPE_BY_BASE_INDEX = {
    **{index: "text" for index in range(1, 5)},
    **{index: "slide" for index in range(5, 9)},
    **{index: "choose" for index in range(9, 13)},
    **{index: "cover" for index in range(13, 17)},
}


@dataclass
class BrowserGeometry:
    screenshot_width_px: int
    screenshot_height_px: int
    viewport_width_css: float
    viewport_height_css: float

    @property
    def x_scale(self) -> float:
        return self.viewport_width_css / self.screenshot_width_px

    @property
    def y_scale(self) -> float:
        return self.viewport_height_css / self.screenshot_height_px


@dataclass
class LaunchTarget:
    html_path: Path
    query_string: str
    description: str


@dataclass
class RunTargetMetadata:
    launch_description: str
    task_group: str
    task_type: str
    task_index: int | None
    variant: int | None


@dataclass
class UsageTotals:
    model_requests: int = 0
    input_tokens: int = 0
    cached_input_tokens: int = 0
    output_tokens: int = 0
    reasoning_tokens: int = 0
    total_tokens: int = 0


@dataclass
class AtomicAction:
    action: str
    x: float
    y: float
    from_x: float
    from_y: float
    to_x: float
    to_y: float
    wheel_direction: str
    wheel_steps: int
    wait_ms: int
    text: str
    clear_first: bool
    key: str
    purpose: str


@dataclass
class AgentDecision:
    decision: str
    screen_state: str
    confidence: float
    plan_note: str
    actions: list[AtomicAction]


@dataclass
class DecisionLog:
    round_number: int
    screenshot_path: str
    screenshot_sha256: str
    decision: dict[str, Any]
    executed_actions: list[dict[str, Any]]
    page_text_excerpt_after_batch: str
    success_detected_after_batch: bool
    screenshot_unchanged_after_batch: bool
    last_batch_likely_failed: bool


class QuietHandler(SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return


def bool_env(name: str, default: bool = False) -> bool:
    value = os.getenv(name, str(default)).strip().lower()
    return value in {"1", "true", "yes", "y", "on"}


def int_env(name: str, default: int) -> int:
    raw = os.getenv(name, str(default)).strip()
    try:
        return int(raw)
    except ValueError as exc:
        raise RuntimeError(
            f"Environment variable {name} must be an integer. Received: {raw}"
        ) from exc


def float_env(name: str, default: float) -> float:
    raw = os.getenv(name, str(default)).strip()
    try:
        return float(raw)
    except ValueError as exc:
        raise RuntimeError(
            f"Environment variable {name} must be a number. Received: {raw}"
        ) from exc


def optional_float_env(name: str) -> float | None:
    raw = os.getenv(name, "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError as exc:
        raise RuntimeError(
            f"Environment variable {name} must be a number when provided. Received: {raw}"
        ) from exc


def results_workbook_path() -> Path:
    raw = os.getenv("AGENT_RESULTS_XLSX", "mi_agent_results.xlsx").strip() or "mi_agent_results.xlsx"
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else ROOT / candidate


def usage_attr(obj: Any, name: str) -> Any:
    if obj is None:
        return None
    if isinstance(obj, dict):
        return obj.get(name)
    return getattr(obj, name, None)


def usage_int(obj: Any, name: str) -> int:
    value = usage_attr(obj, name)
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def capture_response_usage(response: Any, usage_totals: UsageTotals) -> None:
    """
    Collect Gemini usage metadata for the run-level Excel/results summary.

    Google Gen AI responses expose usage_metadata with prompt, candidate,
    thinking, cached-content, and total token counts when available.
    """
    usage_totals.model_requests += 1
    usage = usage_attr(response, "usage_metadata")
    if usage is None:
        return

    usage_totals.input_tokens += usage_int(usage, "prompt_token_count")
    usage_totals.cached_input_tokens += usage_int(usage, "cached_content_token_count")
    usage_totals.output_tokens += usage_int(usage, "candidates_token_count")
    usage_totals.reasoning_tokens += usage_int(usage, "thoughts_token_count")
    usage_totals.total_tokens += usage_int(usage, "total_token_count")

def estimate_run_cost_usd(usage_totals: UsageTotals) -> float | None:
    """
    Optional Gemini cost estimate.

    Provide the price variables in .env to populate estimated_cost_usd:
      GEMINI_INPUT_COST_PER_1M_USD
      GEMINI_CACHED_INPUT_COST_PER_1M_USD
      GEMINI_OUTPUT_COST_PER_1M_USD

    Thinking tokens are included with output billing in the estimate.
    """
    input_price = optional_float_env("GEMINI_INPUT_COST_PER_1M_USD")
    output_price = optional_float_env("GEMINI_OUTPUT_COST_PER_1M_USD")
    cached_input_price = optional_float_env("GEMINI_CACHED_INPUT_COST_PER_1M_USD")

    if input_price is None or output_price is None:
        return None

    cached_price = cached_input_price if cached_input_price is not None else input_price
    cached_tokens = max(0, usage_totals.cached_input_tokens)
    standard_input_tokens = max(0, usage_totals.input_tokens - cached_tokens)
    billable_output_tokens = max(
        0,
        usage_totals.output_tokens + usage_totals.reasoning_tokens,
    )

    cost = (
        standard_input_tokens * input_price
        + cached_tokens * cached_price
        + billable_output_tokens * output_price
    ) / 1_000_000
    return round(cost, 8)

def install_check_press_counter(page: Page) -> None:
    try:
        page.evaluate(
            """
            () => {
              window.__agentCheckPresses = 0;
              document.addEventListener('click', (event) => {
                const rawTarget = event.target;
                const element = rawTarget && rawTarget.closest ? rawTarget : null;
                const button = element ? element.closest('button') : null;
                if (!button) return;
                const label = String(button.innerText || button.textContent || '').trim().toLowerCase();
                if (label.includes('pārbaudīt') || label.includes('verify') || label.includes('check')) {
                  window.__agentCheckPresses += 1;
                }
              }, true);
            }
            """
        )
    except Exception:
        # The agent can still run if metric instrumentation fails.
        return


def read_check_press_count(page: Page) -> int:
    try:
        value = page.evaluate("() => Number(window.__agentCheckPresses || 0)")
        return max(0, int(value or 0))
    except Exception:
        return 0


def variant_for_base_index(task_index: int) -> int:
    return ((task_index - 1) % 4) + 1


def build_run_target_metadata(
    html_path: Path,
    launch_query_string: str,
    launch_description: str,
) -> RunTargetMetadata:
    query = launch_query_string[1:] if launch_query_string.startswith("?") else launch_query_string
    params = parse_qs(query, keep_blank_values=True)
    raw_index = params.get("index", [None])[0]
    task_index: int | None = None
    if raw_index is not None:
        try:
            parsed = int(raw_index)
            if parsed in TASK_TYPE_BY_BASE_INDEX:
                task_index = parsed
        except (TypeError, ValueError):
            task_index = None

    if task_index is not None:
        return RunTargetMetadata(
            launch_description=launch_description,
            task_group="base",
            task_type=TASK_TYPE_BY_BASE_INDEX[task_index],
            task_index=task_index,
            variant=variant_for_base_index(task_index),
        )

    if html_path.name == "puzzle_main_prototype.html":
        return RunTargetMetadata(
            launch_description=launch_description,
            task_group="prototype",
            task_type="shadow_prototype",
            task_index=None,
            variant=None,
        )

    return RunTargetMetadata(
        launch_description=launch_description,
        task_group="unknown",
        task_type="unknown",
        task_index=None,
        variant=None,
    )


def derive_outcome(success: bool, failure_state_detected: bool, decision_round_limit_reached: bool) -> str:
    if success:
        return "success"
    if failure_state_detected:
        return "failure_state_detected"
    if decision_round_limit_reached:
        return "decision_round_limit"
    return "no_confirmed_success"


def start_local_server(port: int, web_root: Path) -> ThreadingHTTPServer:
    os.chdir(web_root)
    server = ThreadingHTTPServer(("127.0.0.1", port), QuietHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server

def parse_gemini_media_resolution(raw: str) -> Any:
    normalized = raw.strip().lower()
    mapping = {
        "default": types.MediaResolution.MEDIA_RESOLUTION_UNSPECIFIED,
        "unspecified": types.MediaResolution.MEDIA_RESOLUTION_UNSPECIFIED,
        "low": types.MediaResolution.MEDIA_RESOLUTION_LOW,
        "medium": types.MediaResolution.MEDIA_RESOLUTION_MEDIUM,
        "high": types.MediaResolution.MEDIA_RESOLUTION_HIGH,
    }
    if normalized not in mapping:
        raise RuntimeError(
            "GEMINI_MEDIA_RESOLUTION must be one of: default, unspecified, low, medium, high."
        )
    return mapping[normalized]

def resolve_gemini_media_resolution_name() -> str:
    """
    Prefer GEMINI_MEDIA_RESOLUTION when provided.
    If it is blank, map IMAGE_DETAIL to the nearest Gemini media setting.
    """
    explicit = os.getenv("GEMINI_MEDIA_RESOLUTION", "").strip()
    if explicit:
        return explicit

    image_detail = os.getenv("IMAGE_DETAIL", "original").strip().lower()
    mapping = {
        "low": "low",
        "high": "high",
        "original": "high",
        "auto": "default",
    }
    return mapping.get(image_detail, "high")

def encode_png_as_data_url(path: Path) -> str:
    encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def screenshot_viewport(page: Page, path: Path) -> BrowserGeometry:
    page.screenshot(path=str(path), full_page=False)
    with Image.open(path) as image:
        screenshot_width, screenshot_height = image.size

    viewport = page.evaluate(
        "() => ({ width: window.innerWidth, height: window.innerHeight })"
    )
    return BrowserGeometry(
        screenshot_width_px=screenshot_width,
        screenshot_height_px=screenshot_height,
        viewport_width_css=float(viewport["width"]),
        viewport_height_css=float(viewport["height"]),
    )


def get_visible_page_text(page: Page, max_chars: int = 1800) -> str:
    try:
        text = page.locator("body").inner_text(timeout=2000).strip()
    except Exception:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text[:max_chars]


def parse_success_texts() -> list[str]:
    raw = os.getenv(
        "SUCCESS_TEXTS",
        "Pārbaude veiksmīga",
    )
    return [item.strip() for item in raw.split("|") if item.strip()]


def parse_failure_texts() -> list[str]:
    raw = os.getenv(
        "FAILURE_TEXTS",
        "Pārbaude neizdevās",
    )
    return [item.strip() for item in raw.split("|") if item.strip()]


def page_contains_any(page: Page, markers: list[str]) -> bool:
    text = get_visible_page_text(page, max_chars=7000).lower()
    return any(marker.lower() in text for marker in markers)


def success_detected(page: Page, success_texts: list[str]) -> bool:
    return page_contains_any(page, success_texts)


def failure_detected(page: Page, failure_texts: list[str]) -> bool:
    return page_contains_any(page, failure_texts)


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def convert_to_css_coordinates(
    geometry: BrowserGeometry,
    x_normalized: float,
    y_normalized: float,
) -> tuple[float, float]:
    """Convert Gemini-style 0..1000 normalized image coordinates to CSS pixels."""
    x_norm = clamp(float(x_normalized), 0.0, 1000.0)
    y_norm = clamp(float(y_normalized), 0.0, 1000.0)
    x_image = (x_norm / 1000.0) * geometry.screenshot_width_px
    y_image = (y_norm / 1000.0) * geometry.screenshot_height_px
    x_css = clamp(x_image * geometry.x_scale, 0, geometry.viewport_width_css - 1)
    y_css = clamp(y_image * geometry.y_scale, 0, geometry.viewport_height_css - 1)
    return x_css, y_css

def html_target_from_relative_path(
    web_root: Path,
    relative_html_path: str,
    query_string: str,
    description: str,
) -> LaunchTarget:
    html_path = (web_root / relative_html_path).resolve()

    if not html_path.exists():
        raise RuntimeError(
            f"The requested launch target is missing: {html_path}. "
            "Make sure the full CAPTCHA folder is next to this agent script."
        )

    if html_path.suffix.lower() != ".html":
        raise RuntimeError(f"The requested launch target is not an .html file: {html_path}")

    return LaunchTarget(
        html_path=html_path,
        query_string=query_string,
        description=description,
    )


def print_cli_usage_and_exit() -> None:
    script_name = Path(sys.argv[0]).name
    print("Usage examples:")
    print(f"  python {script_name} 7")
    print("      Open only the deterministic base task with index 7.")
    print(f"  python {script_name} prototype")
    print("      Open one standalone shadow prototype puzzle.")
    print(f"  python {script_name}")
    print("      Keep the old auto-detection behavior.")
    print("")
    print("Supported base task indexes: 1-16.")
    raise SystemExit(0)


def resolve_cli_launch_target(web_root: Path) -> LaunchTarget | None:
    args = [arg.strip() for arg in sys.argv[1:] if arg.strip()]
    if not args:
        return None

    first = args[0]
    normalized = first.lower()

    if normalized in {"-h", "--help", "help", "/?"}:
        print_cli_usage_and_exit()

    if normalized in {"auto", "detect", "latest"}:
        return None

    if re.fullmatch(r"\d+", first):
        task_index = int(first)
        page_name = BASE_TASK_PAGE_BY_INDEX.get(task_index)
        if not page_name:
            raise RuntimeError(
                f"Unsupported base task index: {task_index}. "
                "Use an integer from 1 to 16."
            )

        return html_target_from_relative_path(
            web_root=web_root,
            relative_html_path=page_name,
            query_string=f"?index={task_index}&order=ordered",
            description=f"base task index {task_index}",
        )

    if normalized in PROTOTYPE_ARGUMENTS:
        return html_target_from_relative_path(
            web_root=web_root,
            relative_html_path="puzzle_main_prototype.html",
            query_string="",
            description="standalone shadow prototype puzzle",
        )

    if normalized.endswith(".html"):
        return html_target_from_relative_path(
            web_root=web_root,
            relative_html_path=first,
            query_string="",
            description=f"explicit HTML file {first}",
        )

    raise RuntimeError(
        f"Unsupported launch argument: {first!r}. "
        "Use 1-16 for a base task, 'prototype' for the shadow prototype, "
        "or run without arguments for the old auto-detection behavior."
    )


def discover_html_file(web_root: Path) -> Path:
    explicit = os.getenv("HTML_FILE", "auto").strip()
    if explicit and explicit.lower() not in {"auto", "detect", "latest"}:
        target = (web_root / explicit).resolve()
        if not target.exists():
            raise RuntimeError(f"HTML_FILE points to a missing file: {target}")
        if target.suffix.lower() != ".html":
            raise RuntimeError(f"HTML_FILE must point to an .html file: {target}")
        return target

    top_level_candidates = sorted(
        [path for path in web_root.glob("*.html") if path.is_file()],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if top_level_candidates:
        candidates = top_level_candidates
    else:
        candidates = []
        for path in web_root.rglob("*.html"):
            if any(part in IGNORED_HTML_DIRS for part in path.parts):
                continue
            if path.is_file():
                candidates.append(path)
        candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)

    if not candidates:
        raise RuntimeError(
            "No HTML file was found. Put one .html file next to this script, "
            "or set HTML_FILE explicitly in .env."
        )

    chosen = candidates[0]
    if len(candidates) == 1:
        print(f"Detected HTML file: {chosen.name}")
    else:
        print("Multiple HTML files detected. Using the most recently modified file:")
        print(f"  -> {chosen.name}")
        print("Other candidates:")
        for path in candidates[1:6]:
            print(f"     {path.name}")
        if len(candidates) > 6:
            print(f"     ... and {len(candidates) - 6} more")
        print("To force a specific file, set HTML_FILE in .env.")
    return chosen.resolve()


def page_url_for_html(
    web_root: Path,
    port: int,
    html_path: Path,
    query_string: str = "",
) -> str:
    try:
        relative = html_path.relative_to(web_root.resolve()).as_posix()
    except ValueError as exc:
        raise RuntimeError(
            "The selected HTML file must be inside the agent folder or one of its subfolders."
        ) from exc

    suffix = query_string.strip()
    if suffix and not suffix.startswith("?"):
        suffix = "?" + suffix

    return f"http://127.0.0.1:{port}/{relative}{suffix}"


def action_signature(action: dict[str, Any]) -> str:
    values = [
        action.get("action", ""),
        round(float(action.get("x", 0.0)), 1),
        round(float(action.get("y", 0.0)), 1),
        round(float(action.get("from_x", 0.0)), 1),
        round(float(action.get("from_y", 0.0)), 1),
        round(float(action.get("to_x", 0.0)), 1),
        round(float(action.get("to_y", 0.0)), 1),
        str(action.get("wheel_direction", "")),
        int(action.get("wheel_steps", 0)),
        str(action.get("text", ""))[:80],
        bool(action.get("clear_first", False)),
        str(action.get("key", "")),
    ]
    return "|".join(map(str, values))


def decision_summary(decision: dict[str, Any]) -> str:
    actions = decision.get("actions", [])
    action_names = [str(action.get("action", "")) for action in actions]
    return (
        f"decision={decision.get('decision')} "
        f"actions={action_names} "
        f"plan_note={decision.get('plan_note', '')!r}"
    )


def last_batch_looked_like_final_check(log: DecisionLog) -> bool:
    if not log.executed_actions:
        return False
    visible_text = log.page_text_excerpt_after_batch.lower()
    check_words = ["pārbaudīt", "verify", "check", "submit", "done", "confirm"]
    clicked_actions = [a for a in log.executed_actions if a.get("action") == "click"]
    if not clicked_actions:
        return False
    return any(word in visible_text for word in check_words)


def history_for_prompt(history: list[DecisionLog], limit: int = 5) -> str:
    if not history:
        return "No prior decision rounds."

    lines: list[str] = []
    for item in history[-limit:]:
        lines.append(
            f"Round {item.round_number}: {decision_summary(item.decision)}; "
            f"success_after={item.success_detected_after_batch}; "
            f"screen_unchanged_after={item.screenshot_unchanged_after_batch}; "
            f"last_batch_likely_failed={item.last_batch_likely_failed}."
        )
    return "\n".join(lines)


def build_reconsideration_hint(
    history: list[DecisionLog],
    low_confidence_threshold: float,
    confidence_drop_delta: float,
) -> str:
    if not history:
        return "No specific reconsideration warning."

    latest = history[-1]
    hints: list[str] = []

    if latest.last_batch_likely_failed:
        hints.append(
            "The previous batch appears to have failed. Re-evaluate whether an earlier choice, "
            "selection, dragged item, or configuration was wrong. Do not merely nudge the same state "
            "unless the screenshot clearly shows that only a small correction is needed."
        )

    if latest.screenshot_unchanged_after_batch:
        hints.append(
            "The previous batch produced no meaningful visible screen change. "
            "Do not repeat the same ineffective actions."
        )

    latest_confidence = clamp(
        float(latest.decision.get("confidence", 1.0)),
        0.0,
        1.0,
    )
    if latest_confidence <= low_confidence_threshold:
        hints.append(
            "Your confidence in the previous decision was low. Step back and reconsider whether "
            "your current interpretation, selected object, target, or overall strategy is correct. "
            "Consider a different approach instead of continuing the same uncertain path."
        )

    if len(history) >= 2:
        last_sig = decision_summary(history[-1].decision)
        prev_sig = decision_summary(history[-2].decision)
        if last_sig == prev_sig:
            hints.append(
                "The last two decision rounds were effectively the same. Choose a materially different strategy."
            )

        previous_confidence = clamp(
            float(history[-2].decision.get("confidence", 1.0)),
            0.0,
            1.0,
        )
        if previous_confidence - latest_confidence >= confidence_drop_delta:
            hints.append(
                "Your confidence dropped noticeably compared with the previous round. Re-check whether "
                "an earlier assumption or action was wrong, and consider an alternative solution path."
            )

    return "\n".join(hints) if hints else "No specific reconsideration warning."


def clean_atomic_action(data: dict[str, Any]) -> AtomicAction:
    required = [
        "action",
        "x",
        "y",
        "from_x",
        "from_y",
        "to_x",
        "to_y",
        "wheel_direction",
        "wheel_steps",
        "wait_ms",
        "text",
        "clear_first",
        "key",
        "purpose",
    ]
    missing = [key for key in required if key not in data]
    if missing:
        raise RuntimeError(f"The model omitted required action fields: {missing}")

    action = str(data["action"])
    if action not in {"click", "drag", "wheel", "wait", "type_text", "press_key"}:
        raise RuntimeError(f"Unsupported atomic action type: {action}")

    wheel_direction = str(data["wheel_direction"])
    if wheel_direction not in {"up", "down", "none"}:
        raise RuntimeError(f"Unsupported wheel direction: {wheel_direction}")

    return AtomicAction(
        action=action,
        x=float(data["x"]),
        y=float(data["y"]),
        from_x=float(data["from_x"]),
        from_y=float(data["from_y"]),
        to_x=float(data["to_x"]),
        to_y=float(data["to_y"]),
        wheel_direction=wheel_direction,
        wheel_steps=max(0, min(120, int(data["wheel_steps"]))),
        wait_ms=max(0, min(5000, int(data["wait_ms"]))),
        text=str(data["text"]),
        clear_first=bool(data["clear_first"]),
        key=str(data["key"]),
        purpose=str(data["purpose"]),
    )


def clean_decision(data: dict[str, Any], max_actions_per_round: int) -> AgentDecision:
    required = ["decision", "screen_state", "confidence", "plan_note", "actions"]
    missing = [key for key in required if key not in data]
    if missing:
        raise RuntimeError(f"The model omitted required decision fields: {missing}")

    decision = str(data["decision"])
    if decision not in {"act", "finish"}:
        raise RuntimeError(f"Unsupported decision type: {decision}")

    raw_actions = data["actions"]
    if not isinstance(raw_actions, list):
        raise RuntimeError("The field actions must be a list.")

    if decision == "finish":
        actions: list[AtomicAction] = []
    else:
        if not raw_actions:
            raise RuntimeError("An act decision must contain at least one action.")
        actions = [
            clean_atomic_action(action)
            for action in raw_actions[:max_actions_per_round]
        ]

    return AgentDecision(
        decision=decision,
        screen_state=str(data["screen_state"]),
        confidence=clamp(float(data["confidence"]), 0.0, 1.0),
        plan_note=str(data["plan_note"]),
        actions=actions,
    )


def ask_model_for_next_decision(
    client: Any,
    model: str,
    screenshot_path: Path,
    history: list[DecisionLog],
    visible_text: str,
    include_visible_text: bool,
    gemini_media_resolution: Any,
    max_actions_per_round: int,
    low_confidence_threshold: float,
    confidence_drop_delta: float,
    usage_totals: UsageTotals,
) -> AgentDecision:
    visible_text_block = (
        f"\nVisible page text, transcribed from the current page:\n{visible_text}\n"
        if include_visible_text and visible_text
        else "\nNo separate page-text transcript is provided. Read any instructions directly from the screenshot.\n"
    )

    prompt = f"""
You are operating a local browser page that contains one visual puzzle task.

Your job is to complete the task using only:
1. the current screenshot,
2. any visible task instructions shown on the page,
3. the short history of your own previous decision rounds below.

Do not assume a task-specific procedure. Infer the task only from the page itself.
The task may require multiple steps.

Return one decision round:
- decision="act": return one or more UI actions to execute now.
- decision="finish": use only when the page visibly indicates the task is already completed.

Supported atomic UI actions:
- click: click a visible UI element, option, tile, button, or selectable item.
- drag: drag from one visible point to another.
- wheel: use the mouse wheel at a visible point. Use wheel_steps to indicate the number of wheel notches. Use wheel_direction="up" or "down".
- type_text: click at x/y, optionally clear the field, then type the visible required text into the input.
- press_key: press a keyboard key such as Enter, Tab, Escape, or Backspace.
- wait: wait briefly only if the UI appears to be updating.

Batching rule:
- You may return multiple actions in one decision round only when they can safely be executed without needing to inspect the screen between them.
- Example of safe batching: selecting several independent matching tiles in a grid, optionally followed by a visible verification button click.
- Another safe batching example: type_text into a clearly visible input field, then click a visible verification button.
- Example of unsafe batching: clicking an object that may create or move a draggable item, then guessing a drag start point before seeing the updated screen. In such a case, output only the first action and re-check next round.
- Prefer batching independent clicks when the page clearly asks for selecting multiple visible items. This reduces unnecessary API rounds.

Coordinate rules:
- All coordinates must be normalized image coordinates from 0 to 1000, relative to the current screenshot.
- x=0 is the far left, x=1000 is the far right, y=0 is the top, and y=1000 is the bottom.
- Do not return raw screenshot pixel coordinates.
- Use x/y for click, wheel, and type_text.
- Use from_x/from_y and to_x/to_y for drag.
- If a coordinate field is unused for the chosen action, set it to 0.
- For type_text, put the exact visible text to type in the text field and set clear_first=true when the input should be cleared before typing.
- For press_key, put the key name in key, for example "Enter".

Quality rules:
- Choose actions that most increase progress.
- If a recent verification/check attempt did not succeed, reassess earlier choices. The problem may be a wrong selection or earlier mistaken decision, not merely a need for small adjustments.
- If a recent batch produced no visible screen change, do not repeat it blindly.
- If a page shows a visible verification/check button, click it only when the visible configuration appears ready.
- If the task asks the user to enter visible text, use type_text instead of giving up.
- If the page likely expects Enter after typing, press_key with key="Enter" is allowed.
- If you are not confident that a sequence is safe to batch, return only one action.

Recent decision history:
{history_for_prompt(history)}

Reconsideration warning:
{build_reconsideration_hint(history, low_confidence_threshold, confidence_drop_delta)}
{visible_text_block}
""".strip()

    schema = {
        "type": "object",
        "properties": {
            "decision": {
                "type": "string",
                "enum": ["act", "finish"],
            },
            "screen_state": {
                "type": "string",
                "description": "Very short factual description of the current visible state.",
            },
            "confidence": {
                "type": "number",
                "minimum": 0,
                "maximum": 1,
            },
            "plan_note": {
                "type": "string",
                "description": "Very short explanation of the decision round.",
            },
            "actions": {
                "type": "array",
                "description": "Atomic UI actions that are safe to execute before another screenshot.",
                "items": {
                    "type": "object",
                    "properties": {
                        "action": {
                            "type": "string",
                            "enum": ["click", "drag", "wheel", "wait", "type_text", "press_key"],
                        },
                        "x": {
                            "type": "number",
                            "description": "Click or wheel x coordinate normalized from 0 to 1000 across the screenshot width. Use 0 when unused.",
                            "minimum": 0,
                            "maximum": 1000,
                        },
                        "y": {
                            "type": "number",
                            "description": "Click or wheel y coordinate normalized from 0 to 1000 across the screenshot height. Use 0 when unused.",
                            "minimum": 0,
                            "maximum": 1000,
                        },
                        "from_x": {
                            "type": "number",
                            "description": "Drag start x coordinate normalized from 0 to 1000 across the screenshot width. Use 0 when unused.",
                            "minimum": 0,
                            "maximum": 1000,
                        },
                        "from_y": {
                            "type": "number",
                            "description": "Drag start y coordinate normalized from 0 to 1000 across the screenshot height. Use 0 when unused.",
                            "minimum": 0,
                            "maximum": 1000,
                        },
                        "to_x": {
                            "type": "number",
                            "description": "Drag end x coordinate normalized from 0 to 1000 across the screenshot width. Use 0 when unused.",
                            "minimum": 0,
                            "maximum": 1000,
                        },
                        "to_y": {
                            "type": "number",
                            "description": "Drag end y coordinate normalized from 0 to 1000 across the screenshot height. Use 0 when unused.",
                            "minimum": 0,
                            "maximum": 1000,
                        },
                        "wheel_direction": {
                            "type": "string",
                            "enum": ["up", "down", "none"],
                            "description": "Use none when action is not wheel.",
                        },
                        "wheel_steps": {
                            "type": "integer",
                            "minimum": 0,
                            "maximum": 120,
                        },
                        "wait_ms": {
                            "type": "integer",
                            "minimum": 0,
                            "maximum": 5000,
                        },
                        "text": {
                            "type": "string",
                            "description": "Text to type for type_text. Use empty string for other actions.",
                        },
                        "clear_first": {
                            "type": "boolean",
                            "description": "For type_text, clear the field before typing when true. Use false for other actions.",
                        },
                        "key": {
                            "type": "string",
                            "description": "Keyboard key for press_key, for example Enter. Use empty string for other actions.",
                        },
                        "purpose": {
                            "type": "string",
                            "description": "Very short purpose of this atomic action.",
                        },
                    },
                    "required": [
                        "action",
                        "x",
                        "y",
                        "from_x",
                        "from_y",
                        "to_x",
                        "to_y",
                        "wheel_direction",
                        "wheel_steps",
                        "wait_ms",
                        "text",
                        "clear_first",
                        "key",
                        "purpose",
                    ],
                },
            },
        },
        "required": [
            "decision",
            "screen_state",
            "confidence",
            "plan_note",
            "actions",
        ],
    }

    image_part = types.Part.from_bytes(
        data=screenshot_path.read_bytes(),
        mime_type="image/png",
    )

    # google-genai's GenerateContentConfig uses response_mime_type +
    # response_schema for structured JSON output. This response_schema path
    # rejects additionalProperties in some SDK/API combinations, so the schema
    # intentionally avoids that OpenAI-style strictness flag.
    config = types.GenerateContentConfig(
        media_resolution=gemini_media_resolution,
        response_mime_type="application/json",
        response_schema=schema,
    )

    response = client.models.generate_content(
        model=model,
        contents=[prompt, image_part],
        config=config,
    )
    capture_response_usage(response, usage_totals)

    response_text = getattr(response, "text", None)
    if not response_text:
        raise RuntimeError("Gemini returned no text output for the decision round.")

    try:
        raw = json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"Gemini did not return valid JSON. Received: {response_text}"
        ) from exc

    return clean_decision(raw, max_actions_per_round=max_actions_per_round)

def execute_atomic_action(
    page: Page,
    geometry: BrowserGeometry,
    action: AtomicAction,
    mouse_drag_steps: int,
    wait_after_action_ms: int,
    wheel_delta_per_step: int,
    wheel_delay_ms: int,
) -> None:
    if action.action == "click":
        x_css, y_css = convert_to_css_coordinates(geometry, action.x, action.y)
        page.mouse.click(x_css, y_css)
        page.wait_for_timeout(wait_after_action_ms)
        return

    if action.action == "drag":
        from_x_css, from_y_css = convert_to_css_coordinates(
            geometry, action.from_x, action.from_y
        )
        to_x_css, to_y_css = convert_to_css_coordinates(
            geometry, action.to_x, action.to_y
        )
        page.mouse.move(from_x_css, from_y_css)
        page.mouse.down()
        page.mouse.move(to_x_css, to_y_css, steps=mouse_drag_steps)
        page.mouse.up()
        page.wait_for_timeout(wait_after_action_ms)
        return

    if action.action == "wheel":
        x_css, y_css = convert_to_css_coordinates(geometry, action.x, action.y)
        page.mouse.move(x_css, y_css)
        direction = -1 if action.wheel_direction == "up" else 1
        steps = max(1, action.wheel_steps)
        for _ in range(steps):
            page.mouse.wheel(0, direction * wheel_delta_per_step)
            if wheel_delay_ms > 0:
                page.wait_for_timeout(wheel_delay_ms)
        page.wait_for_timeout(wait_after_action_ms)
        return

    if action.action == "type_text":
        x_css, y_css = convert_to_css_coordinates(geometry, action.x, action.y)
        page.mouse.click(x_css, y_css)
        if action.clear_first:
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
        if action.text:
            page.keyboard.type(action.text, delay=35)
        page.wait_for_timeout(wait_after_action_ms)
        return

    if action.action == "press_key":
        key = action.key.strip() or "Enter"
        page.keyboard.press(key)
        page.wait_for_timeout(wait_after_action_ms)
        return

    if action.action == "wait":
        page.wait_for_timeout(action.wait_ms or wait_after_action_ms)
        return

    raise RuntimeError(f"Unsupported action: {action.action}")


def execute_batch(
    page: Page,
    geometry: BrowserGeometry,
    decision: AgentDecision,
    mouse_drag_steps: int,
    wait_after_action_ms: int,
    wheel_delta_per_step: int,
    wheel_delay_ms: int,
    success_texts: list[str],
    failure_texts: list[str],
) -> list[dict[str, Any]]:
    executed: list[dict[str, Any]] = []
    for action in decision.actions:
        execute_atomic_action(
            page=page,
            geometry=geometry,
            action=action,
            mouse_drag_steps=mouse_drag_steps,
            wait_after_action_ms=wait_after_action_ms,
            wheel_delta_per_step=wheel_delta_per_step,
            wheel_delay_ms=wheel_delay_ms,
        )
        executed.append(asdict(action))
        if success_detected(page, success_texts):
            break
        if failure_detected(page, failure_texts):
            break
    return executed



def style_header_row(ws: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def ensure_results_runs_sheet(workbook: Any) -> Any:
    if "Runs" in workbook.sheetnames:
        ws = workbook["Runs"]
    else:
        ws = workbook.create_sheet("Runs", 0)

    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        for column, header in enumerate(RESULTS_HEADERS, start=1):
            ws.cell(row=1, column=column).value = header
    elif ws.max_row == 0:
        for column, header in enumerate(RESULTS_HEADERS, start=1):
            ws.cell(row=1, column=column).value = header
    else:
        existing = [ws.cell(row=1, column=col).value for col in range(1, len(RESULTS_HEADERS) + 1)]
        if existing != RESULTS_HEADERS:
            raise RuntimeError(
                "The existing results workbook has unexpected Runs sheet headers. "
                "Rename or move the old workbook, then run again."
            )

    style_header_row(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Y{max(1, ws.max_row)}"

    widths = {
        "A": 23, "B": 21, "C": 20, "D": 30, "E": 14, "F": 18,
        "G": 11, "H": 10, "I": 26, "J": 11, "K": 26, "L": 16,
        "M": 16, "N": 16, "O": 14, "P": 14, "Q": 18, "R": 14,
        "S": 18, "T": 14, "U": 18, "V": 40, "W": 40, "X": 38, "Y": 38,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        row[21].alignment = Alignment(vertical="top", wrap_text=True)  # manual_error_type, V
        row[22].alignment = Alignment(vertical="top", wrap_text=True)  # manual_notes, W

    ws.column_dimensions["V"].width = 44
    ws.column_dimensions["W"].width = 48
    ws.sheet_view.showGridLines = False
    return ws


def ensure_error_type_sheet(workbook: Any) -> Any:
    sheet_name = "ErrorTypeOptions"
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(sheet_name)

    for index, value in enumerate(MANUAL_ERROR_TYPES, start=1):
        ws.cell(row=index, column=1).value = value

    ws.sheet_state = "hidden"
    return ws


def ensure_manual_error_validation(runs_ws: Any, options_ws: Any) -> None:
    has_manual_error_validation = any(
        "V2:V1048576" in str(validation.sqref)
        for validation in runs_ws.data_validations.dataValidation
    )
    if has_manual_error_validation:
        return

    formula = f"'{options_ws.title}'!$A$1:$A${len(MANUAL_ERROR_TYPES)}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.promptTitle = "MI kļūdas veids"
    validation.prompt = "Aizpildi manuāli pēc ekrānattēlu un lēmumu vēstures pārbaudes."
    validation.error = "Izvēlies vienu no piedāvātajiem kļūdas veidiem vai atstāj tukšu."
    validation.errorTitle = "Nederīgs kļūdas veids"
    runs_ws.add_data_validation(validation)
    validation.add("V2:V1048576")


def rebuild_summary_sheet(workbook: Any) -> None:
    if "Summary" in workbook.sheetnames:
        del workbook["Summary"]
    ws = workbook.create_sheet("Summary", 1)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "MI aģenta testu kopsavilkums"
    ws["A1"].font = Font(bold=True, size=16, color="1F1F1F")

    kpis = [
        ("A3", "Testu skaits", "B3", "=COUNTA(Runs!$A$2:$A$1048576)"),
        ("A4", "Veiksmīgi", "B4", '=COUNTIF(Runs!$J:$J,TRUE)'),
        ("A5", "Sekmju īpatsvars", "B5", '=IFERROR(B4/B3,"")'),
        ("A6", "Vidējais laiks (s)", "B6", '=IFERROR(AVERAGE(Runs!$L:$L),"")'),
        ("A7", "Vidējās lēmumu kārtas", "B7", '=IFERROR(AVERAGE(Runs!$M:$M),"")'),
        ("A8", "Modeļa pieprasījumi kopā", "B8", '=SUM(Runs!$N:$N)'),
        ("A9", "Aptuvenās izmaksas (USD)", "B9", '=SUM(Runs!$U:$U)'),
    ]
    for label_cell, label, value_cell, formula in kpis:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = formula

    ws["B5"].number_format = "0.0%"
    ws["B6"].number_format = "0.00"
    ws["B7"].number_format = "0.00"
    ws["B9"].number_format = "$0.000000"

    header_row = 12
    summary_headers = [
        "task_type",
        "tests",
        "successes",
        "success_rate",
        "avg_duration_s",
        "avg_decision_rounds",
        "avg_check_presses",
        "model_requests",
        "estimated_cost_usd",
    ]
    for col, header in enumerate(summary_headers, start=1):
        ws.cell(row=header_row, column=col).value = header

    task_types = ["text", "slide", "choose", "cover", "shadow_prototype", "unknown"]
    for offset, task_type in enumerate(task_types, start=1):
        row = header_row + offset
        ws.cell(row=row, column=1).value = task_type
        ws.cell(row=row, column=2).value = f'=COUNTIF(Runs!$F:$F,A{row})'
        ws.cell(row=row, column=3).value = f'=COUNTIFS(Runs!$F:$F,A{row},Runs!$J:$J,TRUE)'
        ws.cell(row=row, column=4).value = f'=IFERROR(C{row}/B{row},"")'
        ws.cell(row=row, column=5).value = f'=IFERROR(AVERAGEIF(Runs!$F:$F,A{row},Runs!$L:$L),"")'
        ws.cell(row=row, column=6).value = f'=IFERROR(AVERAGEIF(Runs!$F:$F,A{row},Runs!$M:$M),"")'
        ws.cell(row=row, column=7).value = f'=IFERROR(AVERAGEIF(Runs!$F:$F,A{row},Runs!$O:$O),"")'
        ws.cell(row=row, column=8).value = f'=SUMIF(Runs!$F:$F,A{row},Runs!$N:$N)'
        ws.cell(row=row, column=9).value = f'=SUMIF(Runs!$F:$F,A{row},Runs!$U:$U)'
        ws.cell(row=row, column=4).number_format = "0.0%"
        ws.cell(row=row, column=5).number_format = "0.00"
        ws.cell(row=row, column=6).number_format = "0.00"
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "$0.000000"

    style_header_row(ws)
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A13"
    widths = {"A": 22, "B": 12, "C": 14, "D": 15, "E": 18, "F": 22, "G": 20, "H": 18, "I": 20}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def ensure_results_workbook(path: Path) -> Any:
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        default = workbook.active
        default.title = "Runs"

    runs_ws = ensure_results_runs_sheet(workbook)
    options_ws = ensure_error_type_sheet(workbook)
    ensure_manual_error_validation(runs_ws, options_ws)
    rebuild_summary_sheet(workbook)
    return workbook


def append_run_to_results_workbook(path: Path, row_values: list[Any]) -> None:
    workbook = ensure_results_workbook(path)
    runs_ws = workbook["Runs"]
    runs_ws.append(row_values)

    latest_row = runs_ws.max_row
    runs_ws.cell(row=latest_row, column=2).number_format = "yyyy-mm-dd hh:mm:ss"
    runs_ws.cell(row=latest_row, column=12).number_format = "0.00"
    runs_ws.cell(row=latest_row, column=21).number_format = "$0.000000"
    for cell in runs_ws[latest_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=False)
    runs_ws.cell(row=latest_row, column=22).alignment = Alignment(vertical="top", wrap_text=True)
    runs_ws.cell(row=latest_row, column=23).alignment = Alignment(vertical="top", wrap_text=True)
    runs_ws.auto_filter.ref = f"A1:Y{runs_ws.max_row}"

    rebuild_summary_sheet(workbook)

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
    except PermissionError as exc:
        raise RuntimeError(
            f"Could not save the results workbook because it is probably open in Excel: {path}"
        ) from exc


def save_run_summary(
    run_dir: Path,
    html_path: Path,
    model: str,
    logs: list[DecisionLog],
    success: bool,
    metrics: dict[str, Any],
) -> None:
    data = {
        "finished_at": datetime.now().isoformat(timespec="seconds"),
        "html_file": html_path.name,
        "model": model,
        "success": success,
        "decision_round_count": len(logs),
        "metrics": metrics,
        "rounds": [asdict(item) for item in logs],
    }
    (run_dir / "result.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def solve_visual_puzzle() -> None:
    load_dotenv(ROOT / ".env", override=True)

    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY was not found in .env.")

    model = os.getenv("GEMINI_MODEL", "gemini-3.1-pro-preview").strip()
    headless = bool_env("HEADLESS", False)
    port = int_env("SERVER_PORT", 8000)
    viewport_width = int_env("VIEWPORT_WIDTH", 1440)
    viewport_height = int_env("VIEWPORT_HEIGHT", 1100)
    max_decision_rounds = int_env("MAX_DECISION_ROUNDS", 10)
    max_actions_per_round = int_env("MAX_ACTIONS_PER_ROUND", 12)
    low_confidence_threshold = clamp(float_env("LOW_CONFIDENCE_RECONSIDER_THRESHOLD", 0.60), 0.0, 1.0)
    confidence_drop_delta = clamp(float_env("CONFIDENCE_DROP_RECONSIDER_DELTA", 0.20), 0.0, 1.0)
    wait_after_action_ms = int_env("WAIT_AFTER_ACTION_MS", 550)
    mouse_drag_steps = int_env("MOUSE_DRAG_STEPS", 28)
    wheel_delta_per_step = int_env("WHEEL_DELTA_PER_STEP", 100)
    wheel_delay_ms = int_env("WHEEL_DELAY_MS", 35)
    include_visible_text = bool_env("INCLUDE_VISIBLE_PAGE_TEXT", True)
    gemini_media_resolution_name = resolve_gemini_media_resolution_name()
    gemini_media_resolution = parse_gemini_media_resolution(gemini_media_resolution_name)

    success_texts = parse_success_texts()
    failure_texts = parse_failure_texts()

    launch_target = resolve_cli_launch_target(ROOT)
    if launch_target is None:
        html_path = discover_html_file(ROOT)
        launch_query_string = ""
        launch_description = "automatic HTML selection"
    else:
        html_path = launch_target.html_path
        launch_query_string = launch_target.query_string
        launch_description = launch_target.description

    RUNS_DIR.mkdir(exist_ok=True)
    run_dir = RUNS_DIR / datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    run_dir.mkdir(parents=True, exist_ok=False)

    client = genai.Client(api_key=api_key)
    logs: list[DecisionLog] = []
    success = False
    failure_state_detected = False
    usage_totals = UsageTotals()
    workbook_path = results_workbook_path()
    target_metadata = build_run_target_metadata(
        html_path=html_path,
        launch_query_string=launch_query_string,
        launch_description=launch_description,
    )

    server = start_local_server(port, ROOT)

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=headless)
            page = browser.new_page(
                viewport={"width": viewport_width, "height": viewport_height}
            )
            page_url = page_url_for_html(
                ROOT,
                port,
                html_path,
                query_string=launch_query_string,
            )
            page.goto(page_url, wait_until="networkidle")
            page.wait_for_timeout(900)
            install_check_press_counter(page)
            run_started_perf = time.perf_counter()

            print(f"Launch target: {launch_description}")
            print(f"Opened puzzle page: {html_path.name}{launch_query_string}")
            print(f"Model: {model}")
            print(f"Viewport: {viewport_width}x{viewport_height}")
            print(f"Decision-round limit: {max_decision_rounds}")
            print(f"Max actions per round: {max_actions_per_round}")
            print(f"Low-confidence rethink threshold: {low_confidence_threshold:.2f}")
            print(f"Confidence-drop rethink delta: {confidence_drop_delta:.2f}")
            print(f"Visible page text transcript in prompt: {include_visible_text}")
            print(f"Gemini media resolution: {gemini_media_resolution_name}")

            for round_number in range(1, max_decision_rounds + 1):
                if success_detected(page, success_texts):
                    success = True
                    print("Success state is already visible on the page.")
                    break

                if failure_detected(page, failure_texts):
                    failure_state_detected = True
                    print("Failure state is already visible on the page. Stopping this run.")
                    break

                screenshot_path = run_dir / f"round_{round_number:02d}.png"
                geometry = screenshot_viewport(page, screenshot_path)
                screenshot_hash = sha256_file(screenshot_path)
                visible_text = get_visible_page_text(page)

                decision = ask_model_for_next_decision(
                    client=client,
                    model=model,
                    screenshot_path=screenshot_path,
                    history=logs,
                    visible_text=visible_text,
                    include_visible_text=include_visible_text,
                    gemini_media_resolution=gemini_media_resolution,
                    max_actions_per_round=max_actions_per_round,
                    low_confidence_threshold=low_confidence_threshold,
                    confidence_drop_delta=confidence_drop_delta,
                    usage_totals=usage_totals,
                )

                print("=" * 72)
                print(f"Decision round {round_number}/{max_decision_rounds}")
                print(f"Decision: {decision.decision}")
                print(f"Screen state: {decision.screen_state}")
                print(f"Confidence: {decision.confidence:.2f}")
                print(f"Plan note: {decision.plan_note}")
                print(f"Planned actions: {len(decision.actions)}")

                if decision.decision == "finish":
                    success = success_detected(page, success_texts)
                    if success:
                        print("Model selected finish and success text is visible.")
                        break
                    print(
                        "Model selected finish, but no configured success text is visible. "
                        "Continuing."
                    )
                    executed_actions: list[dict[str, Any]] = []
                else:
                    for index, action in enumerate(decision.actions, start=1):
                        print(
                            f"  {index}. {action.action}: {action.purpose}"
                        )
                    executed_actions = execute_batch(
                        page=page,
                        geometry=geometry,
                        decision=decision,
                        mouse_drag_steps=mouse_drag_steps,
                        wait_after_action_ms=wait_after_action_ms,
                        wheel_delta_per_step=wheel_delta_per_step,
                        wheel_delay_ms=wheel_delay_ms,
                        success_texts=success_texts,
                        failure_texts=failure_texts,
                    )

                page_text_after = get_visible_page_text(page)
                success_after_batch = success_detected(page, success_texts)
                failure_after_batch = failure_detected(page, failure_texts)

                after_path = run_dir / f"round_{round_number:02d}_after.png"
                screenshot_viewport(page, after_path)
                after_hash = sha256_file(after_path)
                unchanged_after_batch = after_hash == screenshot_hash

                last_batch_likely_failed = bool(
                    failure_after_batch
                    or (
                        executed_actions
                        and any(action.get("action") == "click" for action in executed_actions)
                        and failure_after_batch
                    )
                )

                logs.append(
                    DecisionLog(
                        round_number=round_number,
                        screenshot_path=str(screenshot_path.relative_to(ROOT)),
                        screenshot_sha256=screenshot_hash,
                        decision={
                            "decision": decision.decision,
                            "screen_state": decision.screen_state,
                            "confidence": decision.confidence,
                            "plan_note": decision.plan_note,
                            "actions": [asdict(action) for action in decision.actions],
                        },
                        executed_actions=executed_actions,
                        page_text_excerpt_after_batch=page_text_after,
                        success_detected_after_batch=success_after_batch,
                        screenshot_unchanged_after_batch=unchanged_after_batch,
                        last_batch_likely_failed=last_batch_likely_failed,
                    )
                )

                if success_after_batch:
                    success = True
                    print("Success state detected after the action batch.")
                    break

                if failure_after_batch:
                    failure_state_detected = True
                    print("Failure state detected after the action batch. Stopping this run.")
                    break

            final_screenshot = run_dir / "final_page.png"
            page.screenshot(path=str(final_screenshot), full_page=True)

            duration_seconds = round(time.perf_counter() - run_started_perf, 3)
            check_presses = read_check_press_count(page)
            decision_round_limit_reached = bool(
                not success
                and not failure_state_detected
                and len(logs) >= max_decision_rounds
            )
            outcome = derive_outcome(
                success=success,
                failure_state_detected=failure_state_detected,
                decision_round_limit_reached=decision_round_limit_reached,
            )
            estimated_cost_usd = estimate_run_cost_usd(usage_totals)
            completed_at = datetime.now().isoformat(timespec="seconds")
            result_json_path = run_dir / "result.json"

            metrics = {
                "run_id": run_dir.name,
                "completed_at": completed_at,
                "launch_target": target_metadata.launch_description,
                "task_group": target_metadata.task_group,
                "task_type": target_metadata.task_type,
                "task_index": target_metadata.task_index,
                "variant": target_metadata.variant,
                "outcome": outcome,
                "duration_seconds": duration_seconds,
                "decision_rounds": len(logs),
                "model_requests": usage_totals.model_requests,
                "check_presses": check_presses,
                "input_tokens": usage_totals.input_tokens,
                "cached_input_tokens": usage_totals.cached_input_tokens,
                "output_tokens": usage_totals.output_tokens,
                "reasoning_tokens": usage_totals.reasoning_tokens,
                "total_tokens": usage_totals.total_tokens,
                "estimated_cost_usd": estimated_cost_usd,
                "failure_state_detected": failure_state_detected,
                "decision_round_limit_reached": decision_round_limit_reached,
            }
            save_run_summary(run_dir, html_path, model, logs, success, metrics)

            result_row = [
                run_dir.name,
                completed_at,
                model,
                target_metadata.launch_description,
                target_metadata.task_group,
                target_metadata.task_type,
                target_metadata.task_index,
                target_metadata.variant,
                html_path.name,
                success,
                outcome,
                duration_seconds,
                len(logs),
                usage_totals.model_requests,
                check_presses,
                usage_totals.input_tokens,
                usage_totals.cached_input_tokens,
                usage_totals.output_tokens,
                usage_totals.reasoning_tokens,
                usage_totals.total_tokens,
                estimated_cost_usd,
                "",
                "",
                str(run_dir),
                str(result_json_path),
            ]
            append_run_to_results_workbook(workbook_path, result_row)

            print("\n" + "=" * 72)
            print("Agent run finished.")
            print(f"Result: {'SUCCESS' if success else 'NO CONFIRMED SUCCESS'}")
            print(f"Outcome: {outcome}")
            print(f"Decision rounds: {len(logs)}")
            print(f"Model requests: {usage_totals.model_requests}")
            print(f"Check button presses: {check_presses}")
            print(f"Duration: {duration_seconds:.3f} s")
            print(f"Run folder: {run_dir}")
            print(f"Updated Excel results: {workbook_path}")
            browser.close()
    finally:
        server.shutdown()
        server.server_close()


if __name__ == "__main__":
    solve_visual_puzzle()

